"""
=============================================================
MODELO DE RISCO DE INADIMPLÊNCIA
Script 3 de 3 — Geração de Scores por Cliente
=============================================================

O que este script faz:
  1. Carrega o modelo treinado (script 02)
  2. Aplica o modelo na base completa de clientes
  3. Gera um score de 0 a 100 e faixa de risco para cada CNPJ
  4. Exporta uma planilha Excel com o resultado final

Saída esperada (colunas na planilha gerada):
  - cnpj               : identificador do cliente
  - score_inadimplencia: score de 0 (baixo risco) a 100 (alto risco)
  - probabilidade_pct  : probabilidade em % de inadimplência
  - faixa_risco        : Baixo / Médio / Alto / Crítico
  - plano              : plano atual do cliente
  - total_cobranças    : total de cobranças na base
  - meses_na_base      : tempo de relacionamento
  - health_score       : health score atual
  - inadimplente_real  : se já é inadimplente (para validação)
=============================================================
"""

import pandas as pd
import numpy as np
import joblib
import warnings
warnings.filterwarnings("ignore")

# ─────────────────────────────────────────────
# CONFIGURAÇÕES
# ─────────────────────────────────────────────

ARQUIVO_BASE     = "base_modelagem.csv"
ARQUIVO_MODELO   = "modelo_inadimplencia.pkl"
ARQUIVO_SCALER   = "scaler.pkl"
ARQUIVO_FEATURES = "features_utilizadas.txt"

ARQUIVO_SAIDA    = "scores_inadimplencia.xlsx"

# Limites das faixas de risco (em probabilidade, 0 a 1)
FAIXAS = {
    "Baixo"   : (0.00, 0.20),   # < 20% de chance
    "Médio"   : (0.20, 0.40),   # 20% a 40%
    "Alto"    : (0.40, 0.65),   # 40% a 65%
    "Crítico" : (0.65, 1.01),   # > 65%
}

# ─────────────────────────────────────────────
# 1. CARREGAMENTO
# ─────────────────────────────────────────────

print("=" * 60)
print("ETAPA 1 — Carregando modelo e dados...")
print("=" * 60)

modelo = joblib.load(ARQUIVO_MODELO)
scaler = joblib.load(ARQUIVO_SCALER)

with open(ARQUIVO_FEATURES) as f:
    features_modelo = [line.strip() for line in f.readlines()]

df = pd.read_csv(ARQUIVO_BASE)
print(f"  Clientes carregados: {len(df):,}")

# ─────────────────────────────────────────────
# 2. APLICAÇÃO DO MODELO
# ─────────────────────────────────────────────

print("\nETAPA 2 — Calculando scores...")

# Seleciona apenas as colunas que o modelo conhece
# (garante que novas dummies de plano não quebrem o modelo)
X = df.reindex(columns=features_modelo, fill_value=0)
X_scaled = scaler.transform(X)

# Probabilidade de inadimplência (0 a 1)
probabilidades = modelo.predict_proba(X_scaled)[:, 1]

# Score de 0 a 100 (invertido: 100 = máximo risco)
scores = (probabilidades * 100).round(1)

# ─────────────────────────────────────────────
# 3. CLASSIFICAÇÃO POR FAIXA DE RISCO
# ─────────────────────────────────────────────

def classificar_faixa(prob):
    for faixa, (low, high) in FAIXAS.items():
        if low <= prob < high:
            return faixa
    return "Crítico"

faixas = [classificar_faixa(p) for p in probabilidades]

# ─────────────────────────────────────────────
# 4. MONTAGEM DO RESULTADO
# ─────────────────────────────────────────────

print("\nETAPA 3 — Montando resultado final...")

resultado = pd.DataFrame({
    "cnpj"              : df["cnpj"],
    "score_inadimplencia": scores,
    "probabilidade_pct" : (probabilidades * 100).round(2),
    "faixa_risco"       : faixas,
    "plano"             : df["plano"] if "plano" in df.columns else "N/A",
    "total_cobranças"   : df["total_cobranças"],
    "meses_na_base"     : df["meses_na_base"],
    "max_dias_atraso"   : df["max_dias_atraso"],
    "media_dias_atraso" : df["media_dias_atraso"].round(1),
    "health_score"      : df["health_score"],
    "inadimplente_real" : df["inadimplente"].map({1: "Sim", 0: "Não"}),
})

# Ordena do maior risco para o menor
resultado = resultado.sort_values("score_inadimplencia", ascending=False)

# ─────────────────────────────────────────────
# 5. RESUMO
# ─────────────────────────────────────────────

print("\n" + "=" * 60)
print("RESUMO DOS SCORES")
print("=" * 60)

for faixa in ["Crítico", "Alto", "Médio", "Baixo"]:
    qtd = (resultado["faixa_risco"] == faixa).sum()
    pct = qtd / len(resultado)
    print(f"  {faixa:<10}: {qtd:>7,} clientes ({pct:.1%})")

print(f"\n  Score médio da base : {resultado['score_inadimplencia'].mean():.1f}")
print(f"  Score mediano       : {resultado['score_inadimplencia'].median():.1f}")

# Clientes de maior atenção (Crítico)
criticos = resultado[resultado["faixa_risco"] == "Crítico"]
if not criticos.empty:
    print(f"\n  Top 10 clientes com maior risco:")
    print(
        criticos.head(10)[["cnpj", "score_inadimplencia", "faixa_risco", "plano", "meses_na_base"]]
        .to_string(index=False)
    )

# ─────────────────────────────────────────────
# 6. EXPORTAÇÃO PARA EXCEL (FORMATADO)
# ─────────────────────────────────────────────

print(f"\nETAPA 4 — Exportando para Excel ({ARQUIVO_SAIDA})...")

with pd.ExcelWriter(ARQUIVO_SAIDA, engine="openpyxl") as writer:

    # ── Aba 1: Todos os clientes com score
    resultado.to_excel(writer, sheet_name="Scores por Cliente", index=False)

    # ── Aba 2: Resumo por faixa de risco
    resumo_faixas = (
        resultado.groupby("faixa_risco")
        .agg(
            clientes        = ("cnpj", "count"),
            score_médio     = ("score_inadimplencia", "mean"),
            inadimplentes   = ("inadimplente_real", lambda x: (x == "Sim").sum()),
        )
        .assign(
            pct_clientes      = lambda x: (x["clientes"] / x["clientes"].sum() * 100).round(1),
            pct_inadimplentes = lambda x: (x["inadimplentes"] / x["clientes"] * 100).round(1),
        )
        .reindex(["Baixo", "Médio", "Alto", "Crítico"])
    )
    resumo_faixas.to_excel(writer, sheet_name="Resumo por Faixa")

    # ── Aba 3: Resumo por plano
    if "plano" in df.columns:
        resumo_plano = (
            resultado.groupby("plano")
            .agg(
                clientes      = ("cnpj", "count"),
                score_médio   = ("score_inadimplencia", "mean"),
                pct_critico   = ("faixa_risco", lambda x: (x == "Crítico").mean() * 100),
                pct_alto      = ("faixa_risco", lambda x: (x == "Alto").mean() * 100),
            )
            .round(1)
            .sort_values("score_médio", ascending=False)
        )
        resumo_plano.to_excel(writer, sheet_name="Resumo por Plano")

    # ── Aba 4: Apenas clientes críticos e de alto risco (lista de ação)
    acao = resultado[resultado["faixa_risco"].isin(["Crítico", "Alto"])]
    acao.to_excel(writer, sheet_name="Lista de Ação", index=False)

# Aplica formatação visual básica (cores nas faixas de risco)
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

CORES_FAIXA = {
    "Crítico": "FF4D4D",
    "Alto"   : "FFA500",
    "Médio"  : "FFD700",
    "Baixo"  : "90EE90",
}

wb = load_workbook(ARQUIVO_SAIDA)

for nome_aba in ["Scores por Cliente", "Lista de Ação"]:
    if nome_aba not in wb.sheetnames:
        continue
    ws = wb[nome_aba]

    # Cabeçalho em negrito
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Encontra coluna de faixa de risco
    col_faixa = None
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value == "faixa_risco":
            col_faixa = col_idx
            break

    if col_faixa:
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            faixa = row[col_faixa - 1].value
            if faixa in CORES_FAIXA:
                fill = PatternFill("solid", fgColor=CORES_FAIXA[faixa])
                row[col_faixa - 1].fill = fill

    # Ajusta largura das colunas
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

wb.save(ARQUIVO_SAIDA)

print(f"\n✓ Exportação concluída: {ARQUIVO_SAIDA}")
print("\n  Abas na planilha:")
print("    → Scores por Cliente  : todos os clientes com score e faixa de risco")
print("    → Resumo por Faixa    : quantos clientes em cada faixa")
print("    → Resumo por Plano    : inadimplência agrupada por produto")
print("    → Lista de Ação       : clientes Críticos e Alto risco (prioridade)")
print("\n  Cores da planilha:")
print("    🔴 Crítico (>65%)  🟠 Alto (40–65%)  🟡 Médio (20–40%)  🟢 Baixo (<20%)")
print("\n✅ Projeto concluído! A planilha está pronta para uso.")
